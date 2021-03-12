from openpyxl import *
from datetime import datetime
import os
import calendar


def get_stok_akhir(vendor_name, wb_data):
    vendor_sheet = wb_data[vendor_name]
    return [stok_akhir.value for stok_akhir in vendor_sheet["L"][12:]]


class NewReport:
    def __init__(self):
        self.today = datetime.today()

    def new_report(self):
        file_dir = "D:/My work/Project/MS/"

        last_month = self.today.replace(month=self.today.month - 1)
        month_range = calendar.monthrange(self.today.year, self.today.month)
        last_month_range = calendar.monthrange(last_month.year, last_month.month)

        wb = load_workbook(f"{file_dir}LAPORAN PENJUALAN VENDOR 1-{last_month_range[1]} "
                           f"{last_month.strftime('%B %Y')}.xlsx")
        wb_data = load_workbook(f"{file_dir}LAPORAN PENJUALAN VENDOR 1-{last_month_range[1]} "
                                f"{last_month.strftime('%B %Y')}.xlsx", data_only=True)

        vendor_names = wb.sheetnames
        stok_akhir_dict = {vendor: get_stok_akhir(vendor, wb_data) for vendor in vendor_names}

        # set stok awal = stok akhir
        for name in vendor_names:
            vendor = wb[name]
            if vendor == wb["SUMMARY"]:
                vendor["I8"] = self.today.strftime("%B")
                vendor["I9"] = self.today.strftime(f"1-{self.today.strftime('%d %B %Y')}")
            else:
                # set periode
                vendor["N8"] = self.today.strftime("=SUMMARY!I8")
                vendor["N9"] = self.today.strftime(f"=SUMMARY!I9")

                for index in range(len(stok_akhir_dict[name])):
                    vendor[f"F{index+13}"] = stok_akhir_dict[name][index]
                    vendor[f"H{index+13}"] = 0
                    vendor[f"N{index+13}"] = ""

        file_name = f"LAPORAN PENJUALAN VENDOR 1-{month_range[1]} {self.today.strftime('%B %Y')}.xlsx"
        wb.save(f"{file_dir}{file_name}")

        os.system(f'start excel.exe "{file_dir}{file_name}"')
