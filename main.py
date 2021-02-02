from openpyxl import *
from datetime import datetime

FILE_DIR = "C:/Users/User/Documents/MS/"

wb = load_workbook(f"{FILE_DIR}LAPORAN PENJUALAN VENDOR 1-30 january.xlsx")
wb_data = load_workbook(f"{FILE_DIR}LAPORAN PENJUALAN VENDOR 1-30 january.xlsx", data_only=True)


def get_stok_akhir(vendor_name):
    vendor_sheet = wb_data[vendor_name]
    return [stok_akhir.value for stok_akhir in vendor_sheet["L"][12:]]


today = datetime.today()
vendor_names = wb.sheetnames

stok_akhir_dict = {vendor: get_stok_akhir(vendor) for vendor in vendor_names}

# set stok awal = stok akhir
for name in vendor_names:
    vendor = wb[name]
    if vendor == wb["SUMMARY"]:
        vendor["I8"] = today.strftime("%B")
        vendor["I9"] = today.strftime(f"1-{today.strftime('%d %B %Y')}")
    else:
        # set periode
        vendor["N8"] = today.strftime("=SUMMARY!I8")
        vendor["N9"] = today.strftime(f"=SUMMARY!I9")

        for index in range(len(stok_akhir_dict[name])):
            vendor[f"F{index+13}"] = stok_akhir_dict[name][index]
            vendor[f"H{index+13}"] = 0
            vendor[f"N{index+13}"] = ""

wb.save(f"{FILE_DIR}LAPORAN PENJUALAN VENDOR 1-{today.strftime('%d %B %Y')}.xlsx")
